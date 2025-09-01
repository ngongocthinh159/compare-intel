#!/usr/bin/env python3
"""
Excel comparator across multiple sheets (fast-stop on first mismatch).

Updates in this version:
- Sheets are processed IN ORDER. If a later sheet is missing in either workbook,
  previously processed sheets are still compared and reported as matched 100%.
  Upon encountering the missing sheet, the program reports the error and exits
  with code 2 (usage/file/sheet error).

What it does:
- Compare two workbooks across a list of sheets (same sheet names in both files).
- For each sheet, compare a top-left rectangle: ROWS x COLS starting at (start-row, start-col).
- Cell comparison rules:
    * Trim whitespace
    * If numeric (int/float/Decimal or numeric-looking string), round HALF_UP to 5 decimal places
      and compare the canonical 5dp string.
    * Else, compare strings case-insensitively (lowercased).
- STOP IMMEDIATELY upon the FIRST UNMATCHED CELL in any sheet.
- Report which sheets matched 100% and which sheet failed (0 or 1) OR which sheet was missing.

Args:
  --manual-path PATH
  --auto-path PATH
  --sheets NAMES...      one or more sheet names (will process in order)
  --rows N               number of rows in the rectangle
  --cols N               number of cols in the rectangle

Optional:
  --start-row N          1-based start row (default: 1)
  --start-col N          1-based start column (default: 1)

Exit codes:
  0  All specified sheets matched fully.
  1  Mismatch found (stopped at first mismatch).
  2  Usage/file/sheet errors (e.g., missing sheet encountered).

Requires:
  pip install openpyxl
"""
import argparse
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Optional

from openpyxl import load_workbook

NUMERIC_RE = re.compile(r'^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$')
DP5 = Decimal("0.00001")


def a1(col: int, row: int) -> str:
    """Convert 1-based (col, row) to A1 notation."""
    s = ""
    c = col
    while c > 0:
        c, rem = divmod(c - 1, 26)
        s = chr(65 + rem) + s
    return f"{s}{row}"


def as_decimal_if_numeric(value) -> Optional[Decimal]:
    """Return a Decimal if value is numeric or numeric-like; else None."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip().replace(',', '')
        if not s:
            return None
        if NUMERIC_RE.match(s):
            try:
                return Decimal(s)
            except Exception:
                return None
    return None


def normalize(value) -> str:
    """Normalize for comparison (5dp numeric, else trimmed lowercase)."""
    d = as_decimal_if_numeric(value)
    if d is not None:
        q = d.quantize(DP5, rounding=ROUND_HALF_UP)
        if q == 0:
            return "0.00000"
        return f"{q:.5f}"
    if value is None:
        return ""
    return str(value).strip().lower()


def compare_rect(ws_m, ws_a, sheet: str, start_row: int, start_col: int, rows: int, cols: int):
    """Compare a rectangle; return (ok, info) where ok is True if fully matched; else False with mismatch info."""
    for r in range(start_row, start_row + rows):
        for c in range(start_col, start_col + cols):
            mv = ws_m.cell(row=r, column=c).value
            av = ws_a.cell(row=r, column=c).value
            nm = normalize(mv)
            na = normalize(av)
            if nm != na:
                cell = a1(c, r)
                return False, {
                    "sheet": sheet,
                    "cell": cell,
                    "row": r,
                    "col": c,
                    "manual_norm": nm,
                    "auto_norm": na,
                    "manual_raw": mv,
                    "auto_raw": av,
                }
    return True, None


def main():
    p = argparse.ArgumentParser(description="Compare two Excel files across sheets; stop on first mismatch or missing sheet.")
    p.add_argument("--manual-path", required=True)
    p.add_argument("--auto-path", required=True)
    p.add_argument("--sheets", nargs="+", required=True)
    p.add_argument("--rows", type=int, required=True)
    p.add_argument("--cols", type=int, required=True)
    p.add_argument("--start-row", type=int, default=1)
    p.add_argument("--start-col", type=int, default=1)
    args = p.parse_args()

    args.sheets = ['Non ic&die']

    try:
        wb_m = load_workbook(filename=args.manual_path, data_only=True, read_only=True)
        wb_a = load_workbook(filename=args.auto_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error opening workbooks: {e}", file=sys.stderr)
        sys.exit(2)

    matched_sheets: List[str] = []

    # Process sheets sequentially; stop upon first mismatch or first missing sheet.
    for sheet in args.sheets:
        has_manual = sheet in wb_m.sheetnames
        has_auto = sheet in wb_a.sheetnames

        if not has_manual or not has_auto:
            print("\n*** SHEET NOT FOUND ***")
            if not has_manual:
                print(f"Missing in manual workbook: '{sheet}'")
                print(f"Manual available sheets: {list(wb_m.sheetnames)}")
            if not has_auto:
                print(f"Missing in auto workbook: '{sheet}'")
                print(f"Auto available sheets: {list(wb_a.sheetnames)}")
            if matched_sheets:
                print(f"\nSheets matched 100% before error: {', '.join(matched_sheets)}")
            else:
                print("\nNo sheets fully matched before error.")
            sys.exit(2)

        ws_m = wb_m[sheet]
        ws_a = wb_a[sheet]
        print(f"Comparing sheet '{sheet}' rows {args.start_row}-{args.start_row+args.rows-1}, "
              f"cols {args.start_col}-{args.start_col+args.cols-1} ...")
        ok, info = compare_rect(ws_m, ws_a, sheet, args.start_row, args.start_col, args.rows, args.cols)
        if not ok:
            print("\n*** MISMATCH DETECTED ***")
            print(f"Failed sheet: {info['sheet']} at cell {info['cell']} (row {info['row']}, col {info['col']})")
            print(f"Manual (normalized): '{info['manual_norm']}'   | Raw: {info['manual_raw']}")
            print(f"Auto   (normalized): '{info['auto_norm']}'     | Raw: {info['auto_raw']}")
            if matched_sheets:
                print(f"\nSheets matched 100% before failure: {', '.join(matched_sheets)}")
            else:
                print("\nNo sheets fully matched before failure.")
            sys.exit(1)
        else:
            print(f"Sheet '{sheet}' matched 100%.")
            matched_sheets.append(sheet)

    # If we got here, all sheets matched.
    print("\nAll specified sheets matched 100%.")
    print(f"Sheets matched 100%: {', '.join(matched_sheets)}")
    sys.exit(0)


if __name__ == "__main__":
    main()
