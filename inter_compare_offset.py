#!/usr/bin/env python3
"""
Sequential row-by-row Excel comparator with offset, with numeric rounding.

Enhancements:
- Before exact match, if a cell is numeric (or a numeric-looking string),
  it is parsed as a Decimal, rounded to 5 decimal places (HALF_UP), and
  then compared using that canonical representation.
- Non-numeric values are still compared as trimmed, lowercased strings.

Other behavior:
- Manual sheet data starts at a given row (default 8).
- Auto sheet data starts at a given row (default 2).
- Start comparing from a zero-based offset: (manual_start+offset) vs (auto_start+offset).
- Compare the first N columns. Stop at first mismatch and print the offset.
- Stop if either side's first N cells are all empty (end-of-data).

Usage:
  python inter_compare_offset.py \
    --manual-path /path/to/manual.xlsx \
    --manual-sheet "ManualSheet" \
    --auto-path /path/to/auto.xlsx \
    --auto-sheet "AutoSheet" \
    --num-cols 7 \
    [--manual-start 8] [--auto-start 2] [--offset 0]

Exit codes:
- 0: compared through to end-of-data with no mismatch
- 1: mismatch encountered
- 2: invalid usage / file/sheet errors

Dependency:
- openpyxl (pip install openpyxl)
"""
import argparse
import sys
import re
from typing import List, Optional
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook


NUMERIC_RE = re.compile(r'^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$')


def as_decimal_if_numeric(value) -> Optional[Decimal]:
    """Return a Decimal if the value is numeric or numeric-like; otherwise None."""
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            # Using str() to preserve user-entered representation for floats.
            return Decimal(str(value))
        except Exception:
            return None
    if isinstance(value, str):
        s = value.strip().replace(',', '')  # drop thousands separators
        if s == "":
            return None
        if NUMERIC_RE.match(s):
            try:
                return Decimal(s)
            except Exception:
                return None
    return None


def normalize(value) -> str:
    """
    Normalize a cell to a canonical comparable string.

    - If numeric: round to 5 decimal places (HALF_UP), format with exactly 5 digits.
      Also collapse negative zero to positive zero.
    - Else: trimmed + lowercased string.
    - None -> empty string.
    """
    d = as_decimal_if_numeric(value)
    if d is not None:
        q = d.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)
        if q == 0:
            return "0.00000"
        return f"{q}"  # Decimal str keeps the quantized places (e.g., 1.23000)
    if value is None:
        return ""
    return str(value).strip().lower()


def read_row(ws, row_index: int, num_cols: int) -> List[str]:
    return [normalize(ws.cell(row=row_index, column=c).value) for c in range(1, num_cols + 1)]


def is_empty_row(values: List[str]) -> bool:
    return all(v == "" for v in values)


def main():
    parser = argparse.ArgumentParser(description="Sequential row comparator with offset; stops at first mismatch. Numeric cells are rounded to 5 decimals before comparison.")
    parser.add_argument("--manual-path", required=True)
    parser.add_argument("--manual-sheet", required=True)
    parser.add_argument("--auto-path", required=True)
    parser.add_argument("--auto-sheet", required=True)
    parser.add_argument("--num-cols", required=True, type=int)
    parser.add_argument("--manual-start", type=int, default=8)
    parser.add_argument("--auto-start", type=int, default=2)
    parser.add_argument("--offset", type=int, default=0, help="Zero-based additional offset from the start rows")
    args = parser.parse_args()

    try:
        wb_m = load_workbook(filename=args.manual_path, data_only=True, read_only=True)
        wb_a = load_workbook(filename=args.auto_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error opening workbooks: {e}", file=sys.stderr)
        sys.exit(2)

    if args.manual_sheet not in wb_m.sheetnames:
        print(f"Manual sheet '{args.manual_sheet}' not found. Available: {wb_m.sheetnames}", file=sys.stderr)
        sys.exit(2)
    if args.auto_sheet not in wb_a.sheetnames:
        print(f"Auto sheet '{args.auto_sheet}' not found. Available: {wb_a.sheetnames}", file=sys.stderr)
        sys.exit(2)

    ws_m = wb_m[args.manual_sheet]
    ws_a = wb_a[args.auto_sheet]

    m_row = args.manual_start + args.offset
    a_row = args.auto_start + args.offset
    current_offset = args.offset
    compared_pairs = 0

    print(f"Starting comparison at offset={current_offset} "
          f"(manual row {m_row} vs auto row {a_row}) over {args.num_cols} columns.")
    print("Note: Numeric values are rounded to 5 decimal places (HALF_UP) before comparison.")

    while True:
        m_vals = read_row(ws_m, m_row, args.num_cols)
        a_vals = read_row(ws_a, a_row, args.num_cols)

        # End condition: if either side's N cells are all empty, stop cleanly.
        if is_empty_row(m_vals):
            print("Success! Reached end-of-data in manual sheet.")
            print(f"Compared {compared_pairs} row pair(s) with no mismatches.")
            print(f"Final offset reached: {current_offset}")
            sys.exit(0)

        # Compare cell-by-cell.
        mismatches = [(idx + 1, mv, av) for idx, (mv, av) in enumerate(zip(m_vals, a_vals)) if mv != av]
        if mismatches:
            print("Mismatch detected!")
            print(f"Offset: {current_offset}")
            print(f"Manual Excel row: {m_row} | Auto Excel row: {a_row}")
            print(f"Columns compared: {args.num_cols}")
            print("Mismatching columns (1-based index, manual_value, auto_value):")
            for col_idx, mv, av in mismatches:
                print(f"  c{col_idx}: '{mv}' vs '{av}'")
            print(f"Matching columns: {args.num_cols - len(mismatches)} / {args.num_cols}")
            sys.exit(1)

        # Otherwise it's a full-row match; advance.
        compared_pairs += 1
        m_row += 1
        a_row += 1
        current_offset += 1

        # Optional progress output every N rows to show we're moving.
        if compared_pairs % 10 == 0:
            print(f"...progress: {compared_pairs} pairs matched (current offset {current_offset})")


if __name__ == "__main__":
    main()
